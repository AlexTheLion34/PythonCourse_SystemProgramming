import openpyxl
from least_square_method.FillTable import *


class Vinification:

    __result = [[]]

    __wb = openpyxl.load_workbook(filename=FillTable.table_file_path)

    __sheet = __wb.active

    __matrix_size = 0

    def __check_length(self):
        j = 5
        current = self.__sheet.cell(row=2, column=j)
        count = 0
        while current.value is not None:
            current = self.__sheet.cell(row=2, column=j)
            count += 1
            j += 1
        self.__matrix_size = count - 1

    def receive_data(self):
        self.__check_length()
        self.__result = [[0] * self.__matrix_size for _ in range(self.__matrix_size)]
        for i in range(len(self.__result)):
            for j in range(len(self.__result)):
                current = self.__sheet.cell(row=i + 2, column=j + 5).value
                if current == -100:
                    if self.__sheet.cell(row=i + 2, column=j + 6).value is not None and \
                            self.__sheet.cell(row=i + 2, column=j + 6).value != 100:
                        self.__result[i][j] = self.__sheet.cell(row=i + 2, column=j + 6).value
                    elif self.__sheet.cell(row=i + 2, column=j + 4).value is not None and \
                            self.__sheet.cell(row=i + 2, column=j + 4).value != 100:
                        self.__result[i][j] = self.__sheet.cell(row=i + 2, column=j + 4).value
                else:
                    self.__result[i][j] = self.__sheet.cell(row=i + 2, column=j + 5).value

        for line in self.__result:
            print(line)
