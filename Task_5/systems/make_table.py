from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.chart import Reference, Series, ScatterChart
import math


def make_table():
    table_file_path = "/Users/AlexTheLion/Desktop/PythonCourse_SystemProgramming/Task_5/systems/systems.xlsx"
    wb = Workbook()
    ws = wb.active
    system_names = ['Линейная', 'Динамическая', 'Любая']
    cells = ['A', 'B', 'C', 'D', 'E', 'F']
    linear = [(5, 10), (10, 20), (20, 40), (40, 80), (80, 160)]
    dynamic = [(10, math.sin(10) + 1), (20, math.sin(20) + 1), (40, math.sin(40) + 1),
               (80, math.sin(80) + 1), (160, math.sin(160) + 1)]
    anything = [(5, math.sin(10) + 1), (10, math.sin(20) + 1), (20, math.sin(40) + 1),
                (40, math.sin(80) + 1), (80, math.sin(160) + 1)]
    ws.merge_cells('A1:B1')
    ws.merge_cells('C1:D1')
    ws.merge_cells('E1:F1')

    counter = 0
    for letter in ['A', 'C', 'E']:
        ws[letter + '1'] = system_names[counter]
        ws[letter + '1'].alignment = Alignment(horizontal='center', vertical='center')
        counter += 1

    counter = 0
    for cell in cells:
        if counter % 2 == 0:
            ws[cell + str(2)] = 'X'
            ws[cell + str(2)].alignment = Alignment(horizontal='center', vertical='center')

        else:
            ws[cell + str(2)] = 'Y'
            ws[cell + str(2)].alignment = Alignment(horizontal='center', vertical='center')
        counter += 1

    for i in range(5):
        ws['A' + str(i + 3)] = (linear[i])[0]
        ws['B' + str(i + 3)] = (linear[i])[1]

    for i in range(5):
        ws['C' + str(i + 3)] = (dynamic[i])[0]
        ws['D' + str(i + 3)] = (dynamic[i])[1]

    for i in range(5):
        ws['E' + str(i + 3)] = (anything[i])[0]
        ws['F' + str(i + 3)] = (anything[i])[1]

    chart_l = ScatterChart()
    chart_l.title = "Linear"
    chart_l.style = 13
    x_values = Reference(ws, min_col=1, min_row=3, max_row=7)
    y_values = Reference(ws, min_col=2, min_row=3, max_row=7)
    series = Series(y_values, x_values, title_from_data=True)
    chart_l.series.append(series)

    chart_d = ScatterChart()
    chart_d.title = "Dynamic"
    chart_d.style = 13
    x_values = Reference(ws, min_col=3, min_row=3, max_row=7)
    y_values = Reference(ws, min_col=4, min_row=3, max_row=7)
    series = Series(y_values, x_values, title_from_data=True)
    chart_d.series.append(series)

    chart_a = ScatterChart()
    chart_a.title = "Anything"
    chart_a.style = 13
    x_values = Reference(ws, min_col=5, min_row=3, max_row=7)
    y_values = Reference(ws, min_col=6, min_row=3, max_row=7)
    series = Series(y_values, x_values, title_from_data=True)
    chart_a.series.append(series)

    ws.add_chart(chart_l, 'A9')
    ws.add_chart(chart_d, 'J9')
    ws.add_chart(chart_a, 'S9')
    wb.save(table_file_path)


make_table()
